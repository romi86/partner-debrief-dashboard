"""
Partner Debrief Intelligence Dashboard - Report Export Module
Generates executive-ready reports in multiple formats
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from datetime import datetime
from typing import Dict, List
import json


class ReportExporter:
    """Exports Partner Debrief data to professional reports."""
    
    def __init__(self):
        self.betterup_blue = "4A90E2"
        self.accent_green = "50C878"
        self.header_gray = "2C3E50"
        self.light_gray = "ECF0F1"
        
    def export_partner_report_excel(self, report_data: Dict, output_path: str):
        """Create comprehensive Excel report for a partner."""
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create sheets
        self._create_executive_summary(wb, report_data)
        self._create_metrics_sheet(wb, report_data)
        self._create_themes_sheet(wb, report_data)
        self._create_insights_sheet(wb, report_data)
        self._create_session_details(wb, report_data)
        
        wb.save(output_path)
        
    def _create_executive_summary(self, wb: Workbook, data: Dict):
        """Create executive summary sheet."""
        ws = wb.create_sheet("Executive Summary")
        
        # Title
        ws['A1'] = "Partner Debrief Intelligence Report"
        ws['A1'].font = Font(size=20, bold=True, color=self.header_gray)
        ws['A2'] = f"Partner: {data['partner_name']}"
        ws['A2'].font = Font(size=16, bold=True, color=self.betterup_blue)
        ws['A3'] = f"Report Generated: {datetime.now().strftime('%B %d, %Y')}"
        ws['A3'].font = Font(size=11, italic=True)
        ws['A4'] = "Created by: Romina Labanca, Coach Community Associate | BetterUp"
        ws['A4'].font = Font(size=10, italic=True, color='666666')
        
        # Key Metrics
        metrics = data['metrics']
        row = 6
        
        ws[f'A{row}'] = "KEY PERFORMANCE INDICATORS"
        ws[f'A{row}'].font = Font(size=14, bold=True, color=self.header_gray)
        ws[f'A{row}'].fill = PatternFill(start_color=self.light_gray, fill_type="solid")
        row += 2
        
        kpi_data = [
            ("Total Survey Responses", metrics.get('total_responses', 0)),
            ("Unique Debrief Sessions", metrics.get('unique_sessions', 0)),
            ("Average Relevance Score", f"{metrics.get('avg_relevance', 0):.2f} / 5.00"),
            ("Average Support Score", f"{metrics.get('avg_support', 0):.2f} / 5.00"),
            ("Average Urgency Score", f"{metrics.get('avg_urgency', 0):.2f} / 5.00"),
        ]
        
        for label, value in kpi_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'] = value
            ws[f'C{row}'].font = Font(size=12, color=self.betterup_blue)
            ws[f'C{row}'].alignment = Alignment(horizontal='right')
            row += 1
        
        # Date range
        if 'date_range' in metrics:
            row += 1
            ws[f'A{row}'] = "Data Period"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'] = f"{metrics['date_range'][0]} to {metrics['date_range'][1]}"
            ws[f'C{row}'].alignment = Alignment(horizontal='right')
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['C'].width = 25
        
    def _create_metrics_sheet(self, wb: Workbook, data: Dict):
        """Create detailed metrics sheet."""
        ws = wb.create_sheet("Detailed Metrics")
        
        # Header
        ws['A1'] = "Coach Satisfaction Metrics"
        ws['A1'].font = Font(size=16, bold=True, color=self.header_gray)
        
        # Time series data
        if not data['time_series'].empty:
            row = 3
            ws[f'A{row}'] = "Session Date"
            ws[f'B{row}'] = "Relevance"
            ws[f'C{row}'] = "Support"
            ws[f'D{row}'] = "Urgency"
            ws[f'E{row}'] = "Response Count"
            
            # Style headers
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{row}'].font = Font(bold=True, color='FFFFFF')
                ws[f'{col}{row}'].fill = PatternFill(start_color=self.betterup_blue, fill_type="solid")
                ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
            
            row += 1
            for _, ts_row in data['time_series'].iterrows():
                ws[f'A{row}'] = ts_row['Date'].strftime('%Y-%m-%d')
                ws[f'B{row}'] = round(ts_row.get('Relevance', 0), 2)
                ws[f'C{row}'] = round(ts_row.get('Support', 0), 2)
                ws[f'D{row}'] = round(ts_row.get('Urgency', 0), 2)
                ws[f'E{row}'] = int(ts_row.get('Response_Count', 0))
                
                # Format cells
                for col in ['B', 'C', 'D']:
                    ws[f'{col}{row}'].number_format = '0.00'
                ws[f'E{row}'].alignment = Alignment(horizontal='center')
                
                row += 1
        
        # Set column widths
        for col, width in [('A', 15), ('B', 12), ('C', 12), ('D', 12), ('E', 15)]:
            ws.column_dimensions[col].width = width
    
    def _create_themes_sheet(self, wb: Workbook, data: Dict):
        """Create themes analysis sheet."""
        ws = wb.create_sheet("Theme Analysis")
        
        # Header
        ws['A1'] = "Coaching Themes & Patterns"
        ws['A1'].font = Font(size=16, bold=True, color=self.header_gray)
        
        themes = data.get('themes', {})
        row = 3
        
        theme_sections = [
            ('organizational_pressures', 'Organizational Pressures'),
            ('leadership_challenges', 'Leadership Challenges'),
            ('implementation_obstacles', 'Implementation Obstacles'),
            ('valuable_takeaways', 'Valuable Takeaways')
        ]
        
        for theme_key, theme_title in theme_sections:
            if theme_key in themes and themes[theme_key]:
                # Section header
                ws[f'A{row}'] = theme_title.upper()
                ws[f'A{row}'].font = Font(size=13, bold=True, color='FFFFFF')
                ws[f'A{row}'].fill = PatternFill(start_color=self.betterup_blue, fill_type="solid")
                ws[f'B{row}'].fill = PatternFill(start_color=self.betterup_blue, fill_type="solid")
                ws.merge_cells(f'A{row}:B{row}')
                row += 1
                
                # Column headers
                ws[f'A{row}'] = "Theme"
                ws[f'B{row}'] = "Frequency"
                for col in ['A', 'B']:
                    ws[f'{col}{row}'].font = Font(bold=True)
                    ws[f'{col}{row}'].fill = PatternFill(start_color=self.light_gray, fill_type="solid")
                row += 1
                
                # Theme data
                for theme, count in themes[theme_key][:10]:
                    ws[f'A{row}'] = theme
                    ws[f'B{row}'] = count
                    ws[f'B{row}'].alignment = Alignment(horizontal='center')
                    row += 1
                
                row += 2
        
        # Set column widths
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 12
    
    def _create_insights_sheet(self, wb: Workbook, data: Dict):
        """Create qualitative insights sheet."""
        ws = wb.create_sheet("Qualitative Insights")
        
        # Header
        ws['A1'] = "Coach Insights & Feedback"
        ws['A1'].font = Font(size=16, bold=True, color=self.header_gray)
        
        insights = data.get('qualitative_insights', {})
        row = 3
        
        for column_name, responses in insights.items():
            if responses:
                # Section header
                ws[f'A{row}'] = self._clean_column_name(column_name)
                ws[f'A{row}'].font = Font(size=12, bold=True, color='FFFFFF')
                ws[f'A{row}'].fill = PatternFill(start_color=self.accent_green, fill_type="solid")
                ws.merge_cells(f'A{row}:B{row}')
                row += 1
                
                # Responses
                for i, response in enumerate(responses[:20], 1):
                    ws[f'A{row}'] = f"{i}."
                    ws[f'A{row}'].alignment = Alignment(vertical='top')
                    ws[f'B{row}'] = str(response)
                    ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                    ws.row_dimensions[row].height = max(30, len(str(response)) // 3)
                    row += 1
                
                row += 2
        
        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 80
    
    def _create_session_details(self, wb: Workbook, data: Dict):
        """Create detailed session data sheet."""
        ws = wb.create_sheet("Session Details")
        
        # Header
        ws['A1'] = "Complete Session Data"
        ws['A1'].font = Font(size=16, bold=True, color=self.header_gray)
        
        if data.get('session_details'):
            # Convert to DataFrame for easier handling
            df = pd.DataFrame(data['session_details'])
            
            # Select key columns
            columns_to_include = [
                'Timestamp', 'Debrief Session Date', 'Coach ID',
                'How relevant was today\'s discussion to your current executive coaching challenges?',
                'How supported do you feel by BetterUp to show up fully in your executive coaching sessions?',
                'How urgent is the primary pressure/challenge you discussed today?',
                'What single organizational pressure is most frequently mentioned by your executives right now?',
                'What leadership challenge or development need keeps coming up across multiple executive sessions?'
            ]
            
            df_subset = df[[col for col in columns_to_include if col in df.columns]]
            
            # Write to sheet starting at row 3
            for r_idx, row in enumerate(dataframe_to_rows(df_subset, index=False, header=True), 3):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # Format header row
                    if r_idx == 3:
                        cell.font = Font(bold=True, color='FFFFFF')
                        cell.fill = PatternFill(start_color=self.betterup_blue, fill_type="solid")
                        cell.alignment = Alignment(wrap_text=True, horizontal='center')
                    else:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Auto-adjust column widths
            for column_cells in ws.columns:
                length = max(len(str(cell.value or '')) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = min(50, max(12, length + 2))
    
    def export_comparison_report_excel(self, comparison_df: pd.DataFrame, 
                                       all_themes: Dict[str, Dict], 
                                       output_path: str):
        """Create cross-partner comparison report."""
        wb = Workbook()
        
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Executive Summary
        ws = wb.create_sheet("Cross-Partner Analysis")
        
        ws['A1'] = "Partner Comparison Report"
        ws['A1'].font = Font(size=20, bold=True, color=self.header_gray)
        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y')}"
        ws['A2'].font = Font(size=11, italic=True)
        ws['A3'] = "Created by: Romina Labanca, Coach Community Associate | BetterUp"
        ws['A3'].font = Font(size=10, italic=True, color='666666')
        
        # Comparison table
        row = 5
        ws[f'A{row}'] = "PARTNER PERFORMANCE METRICS"
        ws[f'A{row}'].font = Font(size=14, bold=True, color=self.header_gray)
        ws[f'A{row}'].fill = PatternFill(start_color=self.light_gray, fill_type="solid")
        row += 2
        
        # Write comparison data
        for r_idx, data_row in enumerate(dataframe_to_rows(comparison_df, index=False, header=True), row):
            for c_idx, value in enumerate(data_row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Format header
                if r_idx == row:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color=self.betterup_blue, fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                else:
                    # Format numeric cells
                    if c_idx > 1 and isinstance(value, (int, float)):
                        cell.number_format = '0.00'
                        cell.alignment = Alignment(horizontal='center')
        
        # Set column widths
        for col_idx, col in enumerate(comparison_df.columns, 1):
            ws.column_dimensions[ws.cell(row=row, column=col_idx).column_letter].width = max(15, len(str(col)) + 2)
        
        # Add themes comparison for each partner
        for partner_name, themes in all_themes.items():
            ws = wb.create_sheet(f"{partner_name[:25]}_Themes")
            
            ws['A1'] = f"{partner_name} - Key Themes"
            ws['A1'].font = Font(size=14, bold=True, color=self.header_gray)
            
            row = 3
            for theme_category, theme_list in themes.items():
                if theme_list:
                    ws[f'A{row}'] = self._clean_column_name(theme_category)
                    ws[f'A{row}'].font = Font(size=12, bold=True, color='FFFFFF')
                    ws[f'A{row}'].fill = PatternFill(start_color=self.accent_green, fill_type="solid")
                    ws.merge_cells(f'A{row}:B{row}')
                    row += 1
                    
                    ws[f'A{row}'] = "Theme"
                    ws[f'B{row}'] = "Count"
                    for col in ['A', 'B']:
                        ws[f'{col}{row}'].font = Font(bold=True)
                        ws[f'{col}{row}'].fill = PatternFill(start_color=self.light_gray, fill_type="solid")
                    row += 1
                    
                    for theme, count in theme_list[:10]:
                        ws[f'A{row}'] = theme
                        ws[f'B{row}'] = count
                        ws[f'B{row}'].alignment = Alignment(horizontal='center')
                        row += 1
                    
                    row += 2
            
            ws.column_dimensions['A'].width = 60
            ws.column_dimensions['B'].width = 12
        
        wb.save(output_path)
    
    def _clean_column_name(self, col_name: str) -> str:
        """Clean column names for display."""
        # Remove underscores and capitalize
        cleaned = col_name.replace('_', ' ').title()
        
        # Handle specific cases
        replacements = {
            'Organizational Pressures': 'Top Organizational Pressures',
            'Leadership Challenges': 'Recurring Leadership Challenges',
            'Implementation Obstacles': 'Implementation Obstacles',
            'Valuable Takeaways': 'Most Valuable Takeaways'
        }
        
        return replacements.get(cleaned, cleaned)
